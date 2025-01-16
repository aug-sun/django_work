from django.contrib import admin
from django.contrib import messages
from birix.sendmail import sendmailclient, sendmailmanager
from birix.models import *
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin
import openpyxl
from datetime import datetime, timedelta
from django.utils.html import format_html
import pytz
from django.urls import reverse, path
from django.utils.safestring import mark_safe



class ContragentsAdmin(LoginRequiredMixin, admin.ModelAdmin):

    actions = ['download_excel',]


    list_display = (
            "ca_name", 
            "ca_shortname",
            "ca_inn",
            "ca_kpp",
            "ca_field_of_activity",
            )
    list_filter = (
            "ca_type",
            "registration_date",
            "ca_field_of_activity",
            "key_manager",
            )
    search_fields = (
            "ca_name",
            "ca_inn",
            "ca_kpp",
            "ca_field_of_activity",
            )

    fieldsets = (
            (None, {
                'fields': (
                    'ca_name',
                    'ca_shortname',
                    'ca_inn',
                    'ca_kpp',
                    'ca_field_of_activity',
                )
            }),
        
        )
    list_per_page = 20

    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Contragents Data"

            # Write headers
            header_row = ["ca_name", "ca_shortname", "ca_inn", "ca_kpp", "ca_field_of_activity"]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for contragent in queryset:
                worksheet.cell(row=row_num, column=1).value = contragent.ca_name
                worksheet.cell(row=row_num, column=2).value = contragent.ca_shortname
                worksheet.cell(row=row_num, column=3).value = contragent.ca_inn
                worksheet.cell(row=row_num, column=4).value = contragent.ca_kpp
                worksheet.cell(row=row_num, column=5).value = contragent.ca_field_of_activity
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=contragents.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response



class LoginUsersAdmin(LoginRequiredMixin,admin.ModelAdmin):

    actions = ['download_excel', 'send_access_mail_manager', 'send_access_mail_client']

    list_display = (
            "login",
            "password",
            "date_create",
            "system",
            "contragent",
            "comment_field",
            "account_status",
            )

    list_filter = (
            "system",
            "date_create",

            )
    search_fields = (
            "client_name",
            "login",
            "comment_field",
            "contragent__ca_name",
    )

    fieldsets = (
            (None, {
                'fields': (
                    'login',
                    'email',
                    'password',
                    'date_create',
                    'system',
                    'contragent',
                    'comment_field',
                    'account_status',
                )
            }),
    )

    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'login',
                    'email',
                    'password',
                    'date_create',
                    'system',
                    'contragent',
                    'comment_field',
                    'account_status',
                )
            })
    )
    autocomplete_fields = ('contragent',)
    list_per_page = 20

# Отправка сообщений с данными для входа по чекбоксам
    def send_access_mail_manager(self, request, queryset):
        for obj in queryset:
            try:
                manager_name = obj.contragent.key_manager
                contragent_name = obj.contragent.ca_name
                manager_email = CaContacts.objects.filter(ca_contact_surname = str(manager_name).split(' ')[0]).first().ca_contact_email
                system_url = obj.system.mon_url
            except Exception as e:
                messages.error(request, f'Ошибка, у контрагента не указан менеджер.{e}.')
            else:
                try:
                #Отправка менеджеру
                    sendmailmanager(manager_email, obj.login, obj.password, contragent_name, system_url, request.user.last_name)
                    obj.account_status = 2
                    obj.save()
                #Отправка создателю
                    sendmailmanager(request.user.email, obj.login, obj.password, contragent_name, system_url, request.user.last_name)
                #Отправка начальству
                    if request.user.username != 'alexandr_master':
                        sendmailmanager('it5@suntel-nn.ru', obj.login, obj.password, contragent_name, system_url, request.user.last_name)
                    messages.success(request, f'Письмо успешно отправлено менеджеру {manager_name} для {contragent_name}.')
                except Exception as e:
                    messages.error(request, f'Ошибка при отправке письма: {e}.')
        return None

# Отправка сообщений с данными для входа по чекбоксам
    def send_access_mail_client(self, request, queryset):
        for obj in queryset:
            try:
                contragent_name = obj.contragent.ca_name
                system_url = obj.system.mon_url
                mon_system_name = obj.system.mon_sys_name
            except Exception as e:
                messages.error(request, f'Ошибка, неправильно заполнена форма клиента (в 1с).{e}.')
            else:
                try:
                #Отправка клиенту
                    sendmailclient(obj.email, obj.login, obj.password, mon_system_name, system_url)
                #Отправка начальству
                    sendmailclient('it5@suntel-nn.ru', obj.login, obj.password, contragent_name, system_url)
                    obj.account_status = 2
                    obj.save()
                    messages.success(request, f'Письмо успешно отправлено для {contragent_name}.')
                except Exception as e:
                    messages.error(request, f'Ошибка при отправке письма: {e}.')
        return None
    





    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "LoginUsers Data"

            # Write headers
            header_row = ["login", "email", "password", "date_create", "system", "contragent", "comment_field", "account_status"]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for loginuser in queryset:
                worksheet.cell(row=row_num, column=1).value = str(loginuser.login)
                worksheet.cell(row=row_num, column=2).value = str(loginuser.email)
                worksheet.cell(row=row_num, column=3).value = str(loginuser.password)
                worksheet.cell(row=row_num, column=4).value = str(loginuser.date_create)
                worksheet.cell(row=row_num, column=5).value = str(loginuser.system)
                worksheet.cell(row=row_num, column=6).value = str(loginuser.contragent)
                worksheet.cell(row=row_num, column=7).value = str(loginuser.comment_field)
                worksheet.cell(row=row_num, column=8).value = str(loginuser.account_status)
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=loginusers.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response



class CaObjectsAdmin(LoginRequiredMixin,admin.ModelAdmin):

    actions = ['download_excel',]
    readonly_fields = ('sys_mon', 'object_name', 'object_status', 'owner_contragent', 'owner_user', 'contragent', 'imei')


    list_display = (
            "sys_mon",
            "object_name",
            "object_status",
            "owner_contragent",
            "owner_user",
            "contragent",
            "imei",
            "get_device",
            "get_sim",
            "sys_mon_object_id",
            "upload_button",
            )

    list_filter = (
            "object_status",
            "sys_mon",
#            "contragent",
#            "contragent",
            )
    search_fields = (
            "object_name",
            "contragent__ca_name",
#            "object_status__status_id",
            "owner_user",
            "owner_contragent",
            "imei",
    )

    fieldsets = (
            (None, {
                'fields': (
                    'sys_mon',
                    'object_name',
                    'object_status',
                    'owner_contragent',
                    'owner_user',
                    'contragent',
                    'imei',

                )
            }),
    )
    list_per_page = 20

    def get_device(self, obj):
        if Devices.objects.filter(device_imei=obj.imei).first():
            if obj.imei == None:
                return "Терминал не найден"
            if obj.imei == Devices.objects.filter(device_imei=obj.imei).first().device_imei:
                return [
                        Devices.objects.filter(device_imei=obj.imei).first().device_serial,
                        Devices.objects.filter(device_imei=obj.imei).first().devices_brand,
                        ]

    def get_sim(self, obj):
        if SimCards.objects.filter(terminal_imei=obj.imei).first():
            if obj.imei == None:
                return "Сим не найден"
                
            if obj.imei == SimCards.objects.filter(terminal_imei=obj.imei).first().terminal_imei:
                return [SimCards.objects.filter(terminal_imei=obj.imei).first().sim_iccid,
                        SimCards.objects.filter(terminal_imei=obj.imei).first().sim_tel_number,
                        SimCards.objects.filter(
                            terminal_imei=obj.imei
                            ).first().sim_cell_operator.name
                        ]

    get_device.short_description = 'Терминал'
    get_sim.short_description = 'Симкарта'


    def upload_button(self, obj):
        return mark_safe(f'<a class="button" href="{reverse("upload_file", args=[obj.id])}">Загрузить</a>')
    upload_button.short_description = 'Загрузить файл'

    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "CaObjects Data"

            # Write headers
            header_row = ["sys_mon", "object_name", "object_status", "owner_contragent", "owner_user", "contragent", "imei"]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for caobject in queryset:
                worksheet.cell(row=row_num, column=1).value = str(caobject.sys_mon)
                worksheet.cell(row=row_num, column=2).value = str(caobject.object_name)
                worksheet.cell(row=row_num, column=3).value = str(caobject.object_status)
                worksheet.cell(row=row_num, column=4).value = str(caobject.owner_contragent)
                worksheet.cell(row=row_num, column=5).value = str(caobject.owner_user)
                worksheet.cell(row=row_num, column=6).value = str(caobject.contragent)
                worksheet.cell(row=row_num, column=7).value = str(caobject.imei)
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=caobjects.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response


class GlobalLogAdmin(LoginRequiredMixin,admin.ModelAdmin):
    list_display = (
            "section_type",
            "get_top_info",
            "get_obj_client",
            "field",
            "get_status_old",
            "get_status_new",
            "change_time",
            "sys_id",
            "action",
            )

    list_filter = (
            "section_type",
            "sys_id",
            "field",
            "change_time",
            "action",
            )
    search_fields = (
            "section_type",
            "edit_id",
            "field",
            "change_time",
            "sys_id",
            "action",
            )
    fieldsets = (
            (None, {
                'fields': (
                    'section_type',
                    'edit_id',
                    'field',
                    'old_value',
                    'new_value',
                    'change_time',
                    'sys_id',
                    'action',
                )
            }),
    )
    list_per_page = 20
    date_hierarchy = 'change_time'


    def get_obj_client(self, obj):
        if obj.section_type == 'object':
            if CaObjects.objects.filter(id=obj.edit_id).first():
                return CaObjects.objects.filter(id=obj.edit_id).first().owner_contragent

        if obj.section_type == '1С_client':
            if Contragents.objects.filter(ca_id=obj.edit_id).first():
                return Contragents.objects.filter(ca_id=obj.edit_id).first().ca_name

    def get_status_old(self, obj):
        if obj.section_type == 'object' and obj.field == 'object_status_id':
            if obj.old_value == "0":
                return 'Не было'
            if obj.old_value == "1":
                return 'Новый не на абонентке'
            if obj.old_value == "2":
                return 'Тестоввый не на абонентке'
            if obj.old_value == "3":
                return 'На абонентке'
            if obj.old_value == "4":
                return 'Ждёт перевода'
            if obj.old_value == "5":
                return 'Приостановлен'
            if obj.old_value == "6":
                return 'Переведённый в другую систему'
            if obj.old_value == "7":
                return 'Деактивирован'
            else:
                return obj.old_value

        if obj.section_type == 'sim_card' and obj.field == 'status':
            if obj.old_value == "0":
                return 'Списана'
            if obj.old_value == "1":
                return 'Активна'
            if obj.old_value == "2":
                return 'Приостановлена'
            if obj.old_value == "3":
                return 'Первичная блокировка'
            if obj.old_value == "4":
                return 'Статус не известен'
            if obj.old_value == "5":
                return 'Сезонная блокировка'
            else:
                return obj.old_value
        else:
            return obj.old_value

    def get_status_new(self, obj):
        if obj.section_type == 'object' and obj.field == 'object_status_id':
            if obj.new_value == "0":
                return 'Не было'
            if obj.new_value == "1":
                return 'Новый не на абонентке'
            if obj.new_value == "2":
                return 'Тестоввый не на абонентке'
            if obj.new_value == "3":
                return 'На абонентке'
            if obj.new_value == "4":
                return 'Ждёт перевода'
            if obj.new_value == "5":
                return 'Приостановлен'
            if obj.new_value == "6":
                return 'Переведённый в другую систему'
            if obj.new_value == "7":
                return 'Деактивирован'
            else:
                return obj.new_value

        if obj.section_type == 'sim_card' and obj.field == 'status':
            if obj.new_value == "0":
                return 'Списана'
            if obj.new_value == "1":
                return 'Активна'
            if obj.new_value == "2":
                return 'Приостановлена'
            if obj.new_value == "3":
                return 'Первичная блокировка'
            if obj.new_value == "4":
                return 'Статус не известен'
            if obj.new_value == "5":
                return 'Сезонная блокировка'
            else:
                return obj.new_value
        else:
            return obj.new_value



    def get_top_info(self, obj):
        info_id = obj.edit_id
        section = obj.section_type

        if section == "sim_card":
            sim = SimCards.objects.filter(sim_id=info_id).first()
            if sim:
                return sim.sim_iccid 

        if section == "object":
            obj = CaObjects.objects.filter(id=info_id).first()
            if obj:
                return obj.object_name



    get_obj_client.short_description = "Контрагент"
    get_status_old.short_description = "Старое значение"
    get_status_new.short_description = "Новое значение"
    get_top_info.short_description = "Детализация"




class SimCardsAdmin(LoginRequiredMixin,admin.ModelAdmin):

    actions = ['download_excel',]

    list_display = (
            "sim_iccid",
            "sim_tel_number",
            "sim_cell_operator",
            "sim_date",
            "contragent",
            'itprogrammer',
            'status',
            'block_start',
            "get_end_date",
            "sim_owner",
            "terminal_imei",
            'get_device',
            )

    list_filter = (
            "sim_cell_operator",
            "sim_owner",
            "sim_date",
            'itprogrammer',
            'status',
            "block_start",
            )
    search_fields = (
            "sim_iccid",
            "sim_tel_number",
            "client_name",
            "sim_cell_operator__name",
            "sim_owner",
            "sim_date",
            "contragent__ca_name",
            "terminal_imei",
            "block_start",
            )
    fieldsets = (
            (None, {
                'fields': (
                    'sim_iccid',
                    'sim_tel_number',
                    'client_name',
                    'sim_cell_operator',
                    'sim_owner',
                    'sim_date',
                    'contragent',
                    "terminal_imei",
                    'itprogrammer',
                    'status',
                )
            }),
    )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'sim_iccid',
                    'sim_tel_number',
#                    'client_name',
                    'sim_cell_operator',
                    'sim_owner',
                    'sim_date',
                    'contragent',
                    "terminal_imei",
                    'itprogrammer',
                    'status',

                )
            })
    )
    autocomplete_fields = (
        'contragent',
    )
    list_per_page = 20
    date_hierarchy = 'sim_date'

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by('-sim_date')

    def get_device(self, obj):
        if Devices.objects.filter(device_imei=obj.terminal_imei).first():
            if obj.terminal_imei == Devices.objects.filter(device_imei=obj.terminal_imei).first().device_imei:
                return Devices.objects.filter(device_imei=obj.terminal_imei).first().device_serial

    def get_end_date(self, obj):
        if obj.block_start:
            # Получаем текущую дату с учетом временной зоны
            current_date = datetime.now(pytz.utc)  # Используем UTC или вашу локальную временную зону
            # Приводим block_start к UTC, если он offset-aware
            if obj.block_start.tzinfo is None:
                block_start = obj.block_start.replace(tzinfo=pytz.utc)  # Присваиваем временную зону
            else:
                block_start = obj.block_start

            # Вычисляем дату окончания блокировки
            end_date = block_start + timedelta(days=180)  # 180 дней = 6 месяцев
            
            # Проверяем, превышает ли дата окончания текущую дату
            if end_date < current_date:
                return format_html('<span style="color: red;">{}</span>', end_date.strftime('%Y-%m-%d'))
            
            return end_date.strftime('%Y-%m-%d')
        
        return None

    get_device.short_description = 'Сер. терм'
    get_end_date.short_description = 'Окончание блокировки'

    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "SimCards Data"

            # Write headers
            header_row = ["sim_iccid", "sim_tel_number", "client_name", "sim_cell_operator", "sim_owner", "sim_date", "contragent", "terminal_imei", "itprogrammer", "status", "block_start"]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for sim in queryset:
                worksheet.cell(row=row_num, column=1).value = str(sim.sim_iccid)
                worksheet.cell(row=row_num, column=2).value = str(sim.sim_tel_number)
                worksheet.cell(row=row_num, column=3).value = str(sim.client_name)
                worksheet.cell(row=row_num, column=4).value = str(sim.sim_cell_operator)
                worksheet.cell(row=row_num, column=5).value = str(sim.sim_owner)
                worksheet.cell(row=row_num, column=6).value = str(sim.sim_date)
                worksheet.cell(row=row_num, column=7).value = str(sim.contragent)
                worksheet.cell(row=row_num, column=8).value = str(sim.terminal_imei)
                worksheet.cell(row=row_num, column=9).value = str(sim.itprogrammer)
                worksheet.cell(row=row_num, column=10).value = str(sim.status)
                worksheet.cell(row=row_num, column=10).value = str(sim.block_start)
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=caobjects.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response

class DevicesAdmin(LoginRequiredMixin,admin.ModelAdmin):

    actions = ['download_excel',]

    list_display = (
            "device_serial",
            "device_imei",
            "device_owner",
#            "client_name",
            "terminal_date",
            "devices_brand",
            "sys_mon",
            "contragent",
            'itprogrammer',
            'get_sim',
            )

    list_filter = (
            "devices_brand",
            "terminal_date",
            'itprogrammer',
            "devices_brand__devices_vendor",
            "sys_mon",
            "device_owner",
            )
    search_fields = (
            "device_serial",
            "device_imei",
            "client_name",
            "name_it",
            "contragent__ca_name",

    )
    fieldsets = (
            (None, {
                'fields': (
                    'device_serial',
                    'device_imei',
                    "device_owner",
                    'terminal_date',
                    'devices_brand',
                    'sys_mon',
                    'contragent',
                    'itprogrammer',
                )
            }),
    )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'device_serial',
                    'device_imei',
                    "device_owner",
                    'terminal_date',
                    'devices_brand',
                    'sys_mon',
                    'contragent__ca_id',
                    'itprogrammer',
                )
            })

    )
    #raw_id_fields = ['contragent']
    autocomplete_fields = (
        'contragent',
    )
    list_per_page = 20
    date_hierarchy = 'terminal_date'

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "devices_brand":
            kwargs["queryset"] = DevicesBrands.objects.all().order_by('name')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('save_form/', self.admin_site.admin_view(self.save_form), name='save_form'),
        ]
        return custom_urls + urls

    def handle_save_form(self, request):
        if request.method == "POST":
            data = request.POST.dict()
            data.pop('device_serial', None)
            data.pop('device_imei', None)
            request.session['form_data'] = data
            return JsonResponse({'success': True})
        return JsonResponse({'success': False})

    def add_view(self, request, form_url='', extra_context=None):
        if request.method == "GET" and 'form_data' in request.session:
            form_data = request.session.pop('form_data')
            form = self.get_form(request)(initial=form_data)
        else:
            form = self.get_form(request)()
        
        return super().add_view(request, form_url, extra_context={'form': form})



    def get_sim(self, obj):
        if SimCards.objects.filter(terminal_imei=obj.device_imei).first():
            if obj.device_imei == SimCards.objects.filter(terminal_imei=obj.device_imei).first().terminal_imei:
                return SimCards.objects.filter(terminal_imei=obj.device_imei).first().sim_iccid

    get_sim.short_description = 'Симкарта на устройстве'
#    list_display_links = ('get_sim',)

    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Terminal Data"

            # Write headers
            header_row = [
                    "device_serial",
                    "device_imei",
                    "client_name",
                    "terminal_date",
                    "devices_brand",
                    "sys_mon",
                    "contragent",
                    'itprogrammer',
                    ]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for sim in queryset:
                worksheet.cell(row=row_num, column=1).value = str(sim.device_serial)
                worksheet.cell(row=row_num, column=2).value = str(sim.device_imei)
                worksheet.cell(row=row_num, column=3).value = str(sim.client_name)
                worksheet.cell(row=row_num, column=4).value = str(sim.terminal_date)
                worksheet.cell(row=row_num, column=5).value = str(sim.devices_brand)
                worksheet.cell(row=row_num, column=6).value = str(sim.sys_mon)
                worksheet.cell(row=row_num, column=7).value = str(sim.contragent)
                worksheet.cell(row=row_num, column=8).value = str(sim.itprogrammer)
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=terminal.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response

class DeviceBrandsAdmin(LoginRequiredMixin,admin.ModelAdmin):
    list_display = (
            "name",
            "devices_vendor",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'name',
                    'devices_vendor',

                )
            })
    )

    list_filter = (
            "devices_vendor",
    )
    search_fields = (
            "name",
            "devices_vendor__vendor_name",
    )

class ContactsAdmin(admin.ModelAdmin):

    actions = ['copy_record']

    def copy_record(self, request, queryset):
        for obj in queryset:
            obj.id = None
            obj.save()

    copy_record.short_description = "Копировать запись"

    list_display = (
            "ca_contact_cell_num",
            "ca_contact_email",
            "ca_contact_name",
            "ca_contact_surname",
            "ca",
            "ca_contact_position",
            )

    fieldsets = (
            (None, {
                'fields': (
                    'ca_contact_cell_num',
                    'ca_contact_email',
                    'ca_contact_name',
                    'ca_contact_surname',
                    'ca',
                    'ca_contact_position',

                )
            }),
    )

    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'ca_contact_cell_num',
                    'ca_contact_email',
                    'ca_contact_name',
                    'ca_contact_surname',
                    'ca',
                    'ca_contact_position',
                )
            })
    )

    search_fields = (
            "ca_contact_cell_num",
            "ca_contact_email"
            "ca_contact_name",
            "ca_contact_surname",
            "ca",
    )
    list_filter = (
            "ca_contact_position",
    )
    autocomplete_fields = (
        'ca',
    )
    
class DevicesCommandAdmin(admin.ModelAdmin):

    actions = ['copy_record']

    def copy_record(self, request, queryset):
        for obj in queryset:
            obj.id = None
            obj.save()

    copy_record.short_description = "Копировать запись"

    list_display = (
            "command",
            "device_brand",
            "method",
            "description",
            )

    list_filter = (
            "device_brand",
            "method",
    )

    search_fields = (
            "command",
            "description",
    )

    fieldsets = (
            (None, {
                'fields': (
                    'command',
                    'device_brand',
                    'method',
                    'description',
                )
            }),
    )

    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'command',
                    'device_brand',
                    'method',
                    'description',
                )
            })
    )


class LogAdmin(admin.ModelAdmin):
    list_display = (
            "action_time",
            "object_id",
            "object_repr",
            "get_change_message",
            "get_change_info",
            "content_type",
            "user",
            )

    def get_change_message(self, obj):
        if obj.action_flag == 1:
            return 'Добавлен новый объект'
        if obj.action_flag == 2:
            return 'Объект изменен'
        if obj.action_flag == 3:
            return 'Объект удален'

    def get_change_info(self, obj):
        message = str(obj.change_message).replace('[', '').replace(']', '').replace('{"changed": {"', "").replace('{"added": {}}', "").replace('"}}', "").replace('fields": "', "")
        clear_message = message.encode('utf-8').decode('unicode_escape')
        return clear_message

    get_change_info.short_description = 'Изменения'

    get_change_message.short_description = 'Действие'


class DeviceVendorAdmin(admin.ModelAdmin):
    list_display = (
            "vendor_name",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'vendor_name',

                )
            })
    )
class MonitoringSystemAdmin(admin.ModelAdmin):
    list_display = (
            "mon_sys_name",
            "mon_sys_ca_obj_price_default",
            "mon_sys_obj_price_suntel",
            "mon_url",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'vendor_name',
                    "mon_sys_ca_obj_price_default",
                    "mon_sys_obj_price_suntel",
                    'mon_url',

                )
            })
    )

class ObjectRetranslatorsAdmin(admin.ModelAdmin):
    list_display = (
            "retranslator_name",
            "retranslator_suntel_price",
            "retranslator_ca_price",
            "retrans_adres",
            "retrans_protocol",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'retranslator_name',
                    "retranslator_suntel_price",
                    "retranslator_ca_price",
                    "retrans_adres",
                    "retrans_protocol",

                )
            })
    )
class GroupObjectRetransAdmin(admin.ModelAdmin):
    list_display = (
            "obj",
            "retr",
            "client_name"

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "obj__id",
                    "retr",

                )
            })
    )
    list_filter = (
            "retr",
            )


    autocomplete_fields = (
        'obj',
    )

    search_fields = (
            "obj__object_name",
            "obj__contragent_id__ca_name"
            )


class ObjectSensorsAdmin(admin.ModelAdmin):
    list_display = (
            "sensor_type",
            "sensor_model",
            "sensor_technology",
            "sensor_connect_type",
            "client",
            "sensor_serial",
            "name_installer",
            "installer_id",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "sensor_type",
                    "sensor_model",
                    "sensor_technology",
                    "sensor_connect_type",
                    "client__ca_id",
                    "sensor_serial",
                    "name_installer",
                    "installer_id",

                )
            })
    )
    list_filter = (
            "sensor_type",
            "sensor_model",
            "sensor_technology",
            "sensor_connect_type",
    )
    search_fields = (
            "sensor_serial",
            "sensor_type",
            "sensor_technology",
    )
    autocomplete_fields = (
        'client',
    )


class WarehouseAdmin(admin.ModelAdmin):
    list_display = (
            "add_date",
            "serial_number",
            "availability",
            "terminal_model",
            "sensor",
            "delivery_date",
            "client",
            "comment",
            "whom_issued",
            "affiliation",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "serial_number",
                    "availability",
                    "terminal_model__id",
                    "sensor__sensor_id",
                    "delivery_date",
                    "client__contragent_id",
                    "comment",
                    "whom_issued",
                    "affiliation",
                )
            })
    )
    list_filter = (
            "affiliation",
            "delivery_date",
            "availability",
            "terminal_model",
            "sensor",
    )
    autocomplete_fields = (
        'terminal_model',
        'sensor',
        'client',
    )

class SensorBrandsAdmin(admin.ModelAdmin):
    list_display = (
            "name",
            "sensor_vendor",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "sensor_vendor__id",
                    "name",

                )
            })
    )
    list_filter = (
            'sensor_vendor__id',
            )

    search_fields = (
            "name",
            "sensor_vendor__vendor_name",
    )


    autocomplete_fields = (
        'sensor_vendor',
    )



class SensorVendorAdmin(admin.ModelAdmin):
    list_display = (
            "name",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "name",

                )
            })
    )
    search_fields = (
            "name",
    )


class DeviceDiagnosicAdmin(admin.ModelAdmin):
    list_display = (
            "device",
            "get_imei",
            "get_klient",
            "programmer",
            "brought",
            "comment",
            "accept_date",
            "transfer_date",
            "whom_tranfer",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "device",
                    "programmer",
                    "brought",
                    "comment",
                    "transfer_date",
                    "whom_tranfer",

                )
            })
    )
    list_filter = (
            'programmer',
            'brought',
            'whom_tranfer',
            )
    search_fields = (
            "comment",
            "device__device_serial",
            "device__contragent__ca_name",
            "device__device_imei",
    )
    autocomplete_fields = (
        'device',
    )
    date_hierarchy = 'transfer_date'


    def get_klient(self, obj):
        return obj.device.contragent.ca_name

    get_klient.short_description = 'Клиент'

    def get_imei(self, obj):
        return obj.device.device_imei

    get_imei.short_description = 'IMEI'

#    readonly_fields = ('accept_date',)


class OnecContractsAdmin(admin.ModelAdmin):
    list_display = (
            "name_contract",
            "contract_number",
            "contract_date",
            "contract_status",
            "organization",
            "counterparty",
            "contract_purpose",
            "type_calculations",
            "category",
            "manager",
            "subdivision",
            "contact_person",
            "detailed_calculations",
            "ok_desk_id",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "name_contract",
                    "contract_number",
                    "contract_date",
                    "contract_status",
                    "organization",
                    "partner",
                    "counterparty",
                    "contract_commencement_date",
                    "contract_expiration_date",
                    "contract_purpose",
                    "type_calculations",
                    "category",
                    "manager",
                    "subdivision",
                    "contact_person",
                    "organization_bank_account",
                    "counterparty_bank_account",
                    "detailed_calculations",
                    "unique_partner_identifier",
                    "unique_counterparty_identifier",
                    "ok_desk_id",
                )
            })
    )
    list_filter = (

            "contract_date",
            "contract_status",
 
            "contract_purpose",
            "type_calculations",
            "category",
            "manager",
            )
    search_fields = (
            "name_contract",
            "contract_number",
            "contract_status",
            "organization",
            "counterparty",
            "manager",
            "subdivision",
            "contact_person",
            "organization_bank_account",
            "counterparty_bank_account",
            "detailed_calculations",
    )


class InfoServObjAdmin(admin.ModelAdmin):
    list_display = (
            "serv_obj_sys_mon",
            "info_obj_serv",
            "subscription_start",
            "subscription_end",
            "tel_num_user",
            "service_counter",
            "stealth_type",
            "monitoring_sys",
            "sys_id_obj",
            "sys_login",
            "sys_password",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "serv_obj_sys_mon",
                    "info_obj_serv",
                    "subscription_start",
                    "subscription_end",
                    "tel_num_user",
                    "service_counter",
                    "stealth_type",
                    "monitoring_sys",
                    "sys_id_obj",
                    "sys_login",
                    "sys_password",
                )
            })
    )
    list_filter = (
                    "info_obj_serv",
                    "subscription_start",
                    "subscription_end",
                    "tel_num_user",
                    "service_counter",
                    "stealth_type",
                    "monitoring_sys",
                    "sys_id_obj",
                    "sys_login",
                    "sys_password",
            )
    search_fields = (
                    "serv_obj_sys_mon",
                    "info_obj_serv",
                    "subscription_start",
                    "subscription_end",
                    "tel_num_user",
                    "service_counter",
                    "stealth_type",
                    "monitoring_sys",
                    "sys_id_obj",
                    "sys_login",
                    "sys_password",
    )
    date_hierarchy = 'subscription_start'
    autocomplete_fields = (
        'serv_obj_sys_mon',
    )


admin.site.register(Contragents, ContragentsAdmin)
admin.site.register(LoginUsers, LoginUsersAdmin)
admin.site.register(GlobalLogging, GlobalLogAdmin)
admin.site.register(CaObjects, CaObjectsAdmin)
admin.site.register(SimCards, SimCardsAdmin)
admin.site.register(Devices, DevicesAdmin)
admin.site.register(DevicesBrands, DeviceBrandsAdmin)
admin.site.register(CaContacts, ContactsAdmin)
admin.site.register(DevicesCommands, DevicesCommandAdmin)
admin.site.register(DjangoAdminLog, LogAdmin)
admin.site.register(DevicesVendor, DeviceVendorAdmin)
admin.site.register(MonitoringSystem, MonitoringSystemAdmin)
admin.site.register(ObjectRetranslators, ObjectRetranslatorsAdmin)
admin.site.register(GroupObjectRetrans, GroupObjectRetransAdmin)
#admin.site.register(ObjectSensors, ObjectSensorsAdmin)
#admin.site.register(EquipmentWarehouse, WarehouseAdmin)
#admin.site.register(SensorBrands, SensorBrandsAdmin)
#admin.site.register(SensorVendor, SensorVendorAdmin)
admin.site.register(DevicesDiagnostics, DeviceDiagnosicAdmin)
admin.site.register(OnecContracts, OnecContractsAdmin)
admin.site.register(InfoServObj, InfoServObjAdmin)
